import openpyxl

def ObtenerInformarcion():

    airbnb= []

    #Cargo la hoja
    libro = openpyxl.load_workbook('Lista_airbnb.xlsx')
    hoja = libro['airbnb']

    for i in range(2, hoja.max_row):
        airbnb.append([hoja.cell(row=i+1, column=2).value, #Link
                      hoja.cell(row=i+1, column=3).value, #Precio
                      (hoja.cell(row=i+1, column=8).value, hoja.cell(row=i+1, column=9).value)]) #Latitud y longitud
        
    return airbnb


# list = ObtenerInformarcion()
# print(list)
